"""
告警中心模块蓝图
提供告警列表、告警详情、告警处置、告警导出等接口
"""
import io
import logging
import os
from datetime import datetime, timezone, timedelta
from flask import Blueprint, request, send_file, send_from_directory, current_app
from flask_jwt_extended import jwt_required, get_jwt_identity, get_jwt
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from app import db, limiter
from app.models.alarm import AlarmEvent
from app.models.user import User
from utils.response import success_response, error_response, paginate_response
from utils.permissions import guard_or_admin_required, admin_required
from core.audit_logger import log_audit

logger = logging.getLogger(__name__)

_CST = timezone(timedelta(hours=8))

alarm_bp = Blueprint('alarm', __name__)

_CAPTURE_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    'data', 'alarm_captures'
)


@alarm_bp.route('/capture/<path:filename>', methods=['GET'])
def serve_capture_image(filename):
    if not filename or '..' in filename or filename.startswith('/'):
        return error_response(message='无效文件名', code=400)
    return send_from_directory(_CAPTURE_DIR, filename)


@alarm_bp.route('/list', methods=['GET'])
@jwt_required()
@limiter.exempt
def get_alarm_list():
    """获取告警列表"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    alarm_type = request.args.get('alarm_type', '')
    alarm_level = request.args.get('alarm_level', '')
    handle_status = request.args.get('handle_status', '')
    start_time = request.args.get('start_time', '')
    end_time = request.args.get('end_time', '')
    
    query = AlarmEvent.query
    
    if alarm_type:
        query = query.filter_by(alarm_type=alarm_type)
    if alarm_level:
        query = query.filter_by(alarm_level=alarm_level)
    if handle_status:
        query = query.filter_by(handle_status=handle_status)
    if start_time:
        query = query.filter(AlarmEvent.alarm_time >= start_time)
    if end_time:
        query = query.filter(AlarmEvent.alarm_time <= end_time)
    
    query = query.order_by(AlarmEvent.alarm_time.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    
    items = []
    for alarm in pagination.items:
        item = alarm.to_dict()
        if alarm.handler_id:
            handler = User.query.get(alarm.handler_id)
            if handler:
                item['handler_name'] = handler.real_name or handler.username
        items.append(item)
    
    return paginate_response(
        items=items,
        total=pagination.total,
        page=pagination.page,
        per_page=pagination.per_page
    )


@alarm_bp.route('/<int:alarm_id>', methods=['GET'])
@jwt_required()
def get_alarm_detail(alarm_id):
    """获取告警详情"""
    alarm = AlarmEvent.query.get(alarm_id)
    if not alarm:
        return error_response(message='告警事件不存在', code=404)
    
    detail = alarm.to_dict()
    if alarm.handler_id:
        handler = User.query.get(alarm.handler_id)
        if handler:
            detail['handler_name'] = handler.real_name or handler.username
    
    return success_response(data=detail)


@alarm_bp.route('/<int:alarm_id>/handle', methods=['PUT'])
@guard_or_admin_required
def handle_alarm(alarm_id):
    """处置告警"""
    alarm = AlarmEvent.query.get(alarm_id)
    if not alarm:
        return error_response(message='告警事件不存在', code=404)
    
    if alarm.handle_status == 'handled':
        return error_response(message='该告警已处置', code=400)
    
    data = request.get_json()
    handle_status = data.get('handle_status', 'handled')
    handle_remark = data.get('handle_remark', '')
    
    if handle_status not in ['handled', 'false_alarm']:
        return error_response(message='处置状态无效', code=400)
    
    current_user_id = int(get_jwt_identity())
    alarm.handle_status = handle_status
    alarm.handler_id = current_user_id
    alarm.handle_time = datetime.now(_CST).isoformat()
    alarm.handle_remark = handle_remark
    
    db.session.commit()
    
    log_audit(
        operation_type='handle_alarm',
        operation_content=f'处置告警: ID={alarm_id}, 状态={handle_status}'
    )
    
    result = alarm.to_dict()
    handler = User.query.get(current_user_id)
    if handler:
        result['handler_name'] = handler.real_name or handler.username
    
    return success_response(data=result, message='处置成功')


@alarm_bp.route('/export', methods=['GET'])
@guard_or_admin_required
def export_alarm_log():
    """导出告警日志"""
    alarm_type = request.args.get('alarm_type', '')
    alarm_level = request.args.get('alarm_level', '')
    handle_status = request.args.get('handle_status', '')
    start_time = request.args.get('start_time', '')
    end_time = request.args.get('end_time', '')
    
    query = AlarmEvent.query
    
    if alarm_type:
        query = query.filter_by(alarm_type=alarm_type)
    if alarm_level:
        query = query.filter_by(alarm_level=alarm_level)
    if handle_status:
        query = query.filter_by(handle_status=handle_status)
    if start_time:
        query = query.filter(AlarmEvent.alarm_time >= start_time)
    if end_time:
        query = query.filter(AlarmEvent.alarm_time <= end_time)
    
    query = query.order_by(AlarmEvent.alarm_time.desc())
    alarms = query.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '告警日志'
    
    headers = ['ID', '告警类型', '告警级别', '告警描述', '告警时间', '处置状态', '处置人', '处置时间', '处置备注']
    ws.append(headers)
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    type_map = {
        'face_alarm': '人脸告警',
        'zone_intrusion': '禁区入侵',
        'behavior_abnormal': '行为异常',
        'danger_alarm': '险情告警'
    }
    level_map = {
        'normal': '一般',
        'warning': '警告',
        'critical': '严重'
    }
    status_map = {
        'pending': '待处置',
        'handled': '已处置',
        'false_alarm': '误报'
    }
    
    for alarm in alarms:
        handler_name = ''
        if alarm.handler_id:
            handler = User.query.get(alarm.handler_id)
            if handler:
                handler_name = handler.real_name or handler.username
        
        row = [
            alarm.id,
            type_map.get(alarm.alarm_type, alarm.alarm_type),
            level_map.get(alarm.alarm_level, alarm.alarm_level),
            alarm.alarm_description,
            alarm.alarm_time,
            status_map.get(alarm.handle_status, alarm.handle_status),
            handler_name,
            alarm.handle_time or '',
            alarm.handle_remark or ''
        ]
        ws.append(row)
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'alarm_log_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    log_audit(
        operation_type='export_alarm',
        operation_content=f'导出告警日志: {len(alarms)}条'
    )
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@alarm_bp.route('/clear', methods=['DELETE'])
@admin_required
def clear_all_alarms():
    """清空所有告警记录"""
    try:
        count = AlarmEvent.query.count()
        AlarmEvent.query.delete()
        db.session.commit()
        log_audit(
            operation_type='clear_alarms',
            operation_content=f'清空所有告警记录: {count}条'
        )
        return success_response(message=f'已清空{count}条告警记录')
    except Exception as e:
        db.session.rollback()
        logger.error(f'Failed to clear alarms: {e}')
        return error_response(message='清空失败', code=500)
